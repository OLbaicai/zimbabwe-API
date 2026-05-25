import json
from pathlib import Path
import openpyxl

TXT_DIR = Path(r'D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\WAY1\txt\金属ID-PROMPT15-OUT')
OUTPUT_XLSX = Path(r'D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\WAY1\xlsx\PROMPT15_ocr_result-jin.xlsx')
HEADERS = ['num', 'file_name', 'id_number', 'surname', 'first_name', 'birth_date', 'gender']

wb = openpyxl.Workbook()
ws = wb.active
for col_idx, header in enumerate(HEADERS, 1):
    ws.cell(row=1, column=col_idx, value=header)

txt_files = sorted(TXT_DIR.glob('*.txt'))
success = fail = 0
for row_idx, txt_path in enumerate(txt_files, 2):
    file_name = txt_path.stem + '.jpg'
    id_number = surname = first_name = birth_date = gender = ''
    try:
        raw = txt_path.read_text(encoding='utf-8').strip()
        data = json.loads(raw)
        id_number  = str(data.get('id_number', '')).strip()
        surname    = str(data.get('surname', '')).strip()
        first_name = str(data.get('first_name', '')).strip()
        birth_date = str(data.get('birth_date', '')).strip()
        gender     = str(data.get('gender', '')).strip()
        success += 1
    except Exception as e:
        fail += 1
    ws.cell(row=row_idx, column=1, value='')
    ws.cell(row=row_idx, column=2, value=file_name)
    ws.cell(row=row_idx, column=3, value=id_number)
    ws.cell(row=row_idx, column=4, value=surname)
    ws.cell(row=row_idx, column=5, value=first_name)
    ws.cell(row=row_idx, column=6, value=birth_date)
    ws.cell(row=row_idx, column=7, value=gender)

wb.save(OUTPUT_XLSX)
print(f'TXT->XLSX: success={success} fail={fail} total={len(txt_files)}')
print(f'Saved: {OUTPUT_XLSX}')
