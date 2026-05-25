"""change-xlsx-to-json.py

将 WAY2/deal-data_ocr_result-jin.xlsx 的每一行转为单独的 JSON txt 文件。

输出：WAY2/金属ID-JSON/<image_stem>.txt
每个 txt 内容只包含一个 JSON 对象，且仅包含五个字段：
id_number, surname, first_name, birth_date, gender

字段为空则输出空字符串，不输出其它内容。
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl


XLSX_PATH = Path(r"D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\金属ID标注\my_idcard_zimbabwe_label-jin.xlsx")
OUT_DIR = Path(r"D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\WAY2\金属ID-JSON")

FIELDS = ("id_number", "surname", "first_name", "birth_date", "gender")


def _as_str(value) -> str:
	if value is None:
		return ""
	return str(value).strip()


def _normalize_image_stem(file_name: str) -> str:
	name = file_name.strip()
	if not name:
		return ""
	p = Path(name)
	# 常见：03041586W03.jpg -> 03041586W03
	return p.stem if p.suffix else p.name


def main() -> None:
	if not XLSX_PATH.exists():
		raise FileNotFoundError(f"xlsx 不存在: {XLSX_PATH}")

	OUT_DIR.mkdir(parents=True, exist_ok=True)

	wb = openpyxl.load_workbook(XLSX_PATH)
	ws = wb.active

	# 读取表头，定位列
	header_row = 1
	col_index: dict[str, int] = {}
	for col in range(1, ws.max_column + 1):
		key = _as_str(ws.cell(row=header_row, column=col).value)
		if key:
			col_index[key] = col

	if "file_name" not in col_index:
		raise RuntimeError("xlsx 表头缺少 file_name 列")

	for field in FIELDS:
		if field not in col_index:
			raise RuntimeError(f"xlsx 表头缺少 {field} 列")

	# 逐行写出
	for row in range(2, ws.max_row + 1):
		file_name = _as_str(ws.cell(row=row, column=col_index["file_name"]).value)
		stem = _normalize_image_stem(file_name)
		if not stem:
			continue

		payload = {field: _as_str(ws.cell(row=row, column=col_index[field]).value) for field in FIELDS}
		out_path = OUT_DIR / f"{stem}.txt"
		out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
	main()

