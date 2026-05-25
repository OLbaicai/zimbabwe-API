"""
change_format.py
读取 金属ID-OUT/ 下的 OCR 结果 .txt 文件（JSON格式），
按照 my_idcard_zimbabwe_label-jin.xlsx 的列结构写入新的 .xlsx 文件，
方便与真实标签做对比。
"""

import json
from pathlib import Path

import openpyxl

# ── 路径配置 ──────────────────────────────────────────────
# 读取 WAY2 结构化输出目录（每个 txt 是一个 JSON）
TXT_DIR = Path(r"D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\WAY2\金属ID-MAIN-DEAL-OUT")
# 写到 WAY2 下，供 same-evaluate.py 直接评估
OUTPUT_XLSX = Path(r"D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\WAY2\deal-data_ocr_result-jin.xlsx")

# ── 表头（与 my_idcard_zimbabwe_label-jin.xlsx 一致）─────
HEADERS = ["num", "file_name", "id_number", "surname", "first_name", "birth_date", "gender"]


def main() -> None:
    if not TXT_DIR.exists():
        print(f"[错误] 目录不存在: {TXT_DIR}")
        return

    # 收集所有 .txt 文件，按文件名排序
    txt_files = sorted(p for p in TXT_DIR.iterdir() if p.suffix.lower() == ".txt")
    if not txt_files:
        print(f"[警告] 目录中未找到 .txt 文件: {TXT_DIR}")
        return

    print(f"找到 {len(txt_files)} 个 txt 文件")

    # 创建新工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR-Results"

    # 写表头
    for col_idx, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    success = 0
    fail = 0

    # 逐文件解析写入
    for row_idx, txt_path in enumerate(txt_files, 2):
        # B列：文件名 = txt 的 stem + .jpg
        file_name = txt_path.stem + ".jpg"

        # 默认空值
        id_number = surname = first_name = birth_date = gender = ""

        try:
            raw = txt_path.read_text(encoding="utf-8").strip()
            # 如果模型返回了 markdown 代码块，去掉 ```json 和 ```
            if raw.startswith("```"):
                lines = raw.split("\n")
                # 去掉第一行 ```json 和最后一行 ```
                if lines[-1].strip() == "```":
                    lines = lines[1:-1]
                else:
                    lines = lines[1:]
                raw = "\n".join(lines)

            data = json.loads(raw)
            id_number  = str(data.get("id_number", "")).strip()
            surname    = str(data.get("surname", "")).strip()
            first_name = str(data.get("first_name", "")).strip()
            birth_date = str(data.get("birth_date", "")).strip()
            gender     = str(data.get("gender", "")).strip()
            success += 1

        except json.JSONDecodeError as e:
            print(f"  [跳过] JSON 解析失败: {txt_path.name} — {e}")
            fail += 1
            # 解析失败的也写入，字段为空但文件名保留
        except Exception as e:
            print(f"  [跳过] 读取失败: {txt_path.name} — {e}")
            fail += 1

        # 写入行：A=空, B=文件名, C~G=解析结果
        row_data = ["", file_name, id_number, surname, first_name, birth_date, gender]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 保存
    wb.save(OUTPUT_XLSX)
    print(f"\n写入完成  成功: {success}  失败: {fail}  共: {len(txt_files)}")
    print(f"输出文件: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
