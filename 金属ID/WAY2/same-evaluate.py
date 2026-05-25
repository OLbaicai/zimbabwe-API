"""
evaluate.py
对比 OCR 识别结果 Excel 与真实标签 Excel，输出：
  - 每个字段的准确率
  - 总体准确率
  - 具体错误明细
"""

from pathlib import Path

import openpyxl

# ── 路径配置 ──────────────────────────────────────────────────
OCR_RESULT_XLSX = Path(r"D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\WAY2\deal-data_ocr_result-jin.xlsx")
GROUND_TRUTH_XLSX = Path(r"D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\金属ID标注\my_idcard_zimbabwe_label-jin.xlsx")
REPORT_TXT = Path(r"D:\aaa_my_iwhalecloud\VScode_AI\HunyuanOCR\津巴布韦API开发测试\WAY2\report\deal-data_evaluate_report.txt")

# 需要对比的字段（列索引，1-based）
COMPARE_FIELDS = {3: "id_number", 4: "surname", 5: "first_name", 6: "birth_date", 7: "gender"}


def _load_sheet(path: Path):
    """读取 xlsx，返回 {file_name: {col_idx: value}} 字典。"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    data = {}
    for row in range(2, ws.max_row + 1):
        file_name = ws.cell(row=row, column=2).value
        if file_name is None:
            continue
        file_name = str(file_name).strip()
        row_data = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            row_data[col] = str(val).strip() if val is not None else ""
        data[file_name] = row_data
    return data


def _normalize_date(date_str: str) -> str:
    """统一日期格式：将 D.M.YY、D/M/YY 等统一为 DD.MM.YYYY（用点分隔，与标签格式对齐）。"""
    if not date_str:
        return ""
    # 替换斜杠为点
    s = date_str.replace("/", ".").replace("-", ".").replace(" ", "")
    parts = s.split(".")
    if len(parts) != 3:
        return date_str.strip()
    day, month, year = parts
    # 补零：日、月补为两位
    day = day.zfill(2)
    month = month.zfill(2)
    # 年份补全为四位
    if len(year) == 2:
        year = "19" + year if int(year) > 30 else "20" + year
    return f"{day}.{month}.{year}"


def _normalize_gender(g: str) -> str:
    """统一性别：M/male -> male, F/female -> female"""
    g = g.strip().lower()
    if g in ("m", "male"):
        return "male"
    if g in ("f", "female"):
        return "female"
    return g


def main() -> None:
    if not OCR_RESULT_XLSX.exists():
        print(f"[错误] OCR 结果文件不存在: {OCR_RESULT_XLSX}")
        return
    if not GROUND_TRUTH_XLSX.exists():
        print(f"[错误] 标签文件不存在: {GROUND_TRUTH_XLSX}")
        return

    print("加载 OCR 结果...")
    ocr_data = _load_sheet(OCR_RESULT_XLSX)
    print(f"  OCR 记录数: {len(ocr_data)}")

    print("加载真实标签...")
    gt_data = _load_sheet(GROUND_TRUTH_XLSX)
    print(f"  标签记录数: {len(gt_data)}")

    # ── 匹配 ──────────────────────────────────────────────────
    common_files = sorted(set(ocr_data.keys()) & set(gt_data.keys()))
    only_ocr = sorted(set(ocr_data.keys()) - set(gt_data.keys()))
    only_gt  = sorted(set(gt_data.keys()) - set(ocr_data.keys()))

    print(f"\n匹配到的共同文件: {len(common_files)}")
    if only_ocr:
        print(f"仅 OCR 中有（标签无）: {len(only_ocr)} 个")
    if only_gt:
        print(f"仅标签中有（OCR 无）: {len(only_gt)} 个")

    if not common_files:
        print("[错误] 没有共同文件可比较")
        return

    REPORT_TXT.parent.mkdir(parents=True, exist_ok=True)

    # ── 逐字段比较 ────────────────────────────────────────────
    field_correct = {f: 0 for f in COMPARE_FIELDS.values()}
    field_total = len(common_files)

    errors: list[dict] = []  # 每条错误: {file_name, field, label, ocr}

    for file_name in common_files:
        gt_row = gt_data[file_name]
        ocr_row = ocr_data[file_name]
        for col_idx, field_name in COMPARE_FIELDS.items():
            gt_val = gt_row.get(col_idx, "")
            ocr_val = ocr_row.get(col_idx, "")

            # 日期/性别做归一化后再比较
            if field_name == "birth_date":
                gt_norm = _normalize_date(gt_val)
                ocr_norm = _normalize_date(ocr_val)
            elif field_name == "gender":
                gt_norm = _normalize_gender(gt_val)
                ocr_norm = _normalize_gender(ocr_val)
            else:
                gt_norm = gt_val.strip()
                ocr_norm = ocr_val.strip()

            if gt_norm == ocr_norm:
                field_correct[field_name] += 1
            else:
                errors.append({
                    "file_name": file_name,
                    "field": field_name,
                    "label": gt_val,
                    "ocr": ocr_val,
                })

    # ── 终端报告 ──────────────────────────────────────────────
    print(f"\n{'=' * 65}")
    print(f"  字段准确率（共 {field_total} 条）")
    print(f"{'=' * 65}")
    total_correct = 0
    total_fields = field_total * len(COMPARE_FIELDS)
    for field_name in COMPARE_FIELDS.values():
        correct = field_correct[field_name]
        rate = correct / field_total * 100
        bar = "#" * int(rate / 5) + "-" * (20 - int(rate / 5))
        print(f"  {field_name:<14s}  {correct:>4d}/{field_total}  [{bar}]  {rate:5.1f}%")
        total_correct += correct

    overall_rate = total_correct / total_fields * 100
    print(f"  {'─' * 55}")
    print(f"  {'总体准确率':<14s}  {total_correct:>4d}/{total_fields}  [{'#' * int(overall_rate / 5) + '-' * (20 - int(overall_rate / 5))}]  {overall_rate:5.1f}%")
    print(f"{'=' * 65}")

    error_count = len(errors)

    # ── 构建 txt 报告 ─────────────────────────────────────────
    lines: list[str] = []
    sep = "=" * 65
    sep2 = "-" * 55

    lines.append(sep)
    lines.append(f"  字段准确率（共 {field_total} 条）")
    lines.append(sep)
    total_correct = 0
    total_fields = field_total * len(COMPARE_FIELDS)
    for field_name in COMPARE_FIELDS.values():
        correct = field_correct[field_name]
        rate = correct / field_total * 100
        bar = "#" * int(rate / 5) + "-" * (20 - int(rate / 5))
        lines.append(f"  {field_name:<14s}  {correct:>4d}/{field_total}  [{bar}]  {rate:5.1f}%")
        total_correct += correct

    overall_rate = total_correct / total_fields * 100
    lines.append(f"  {sep2}")
    lines.append(f"  {'总体准确率':<14s}  {total_correct:>4d}/{total_fields}  [{'#' * int(overall_rate / 5) + '-' * (20 - int(overall_rate / 5))}]  {overall_rate:5.1f}%")
    lines.append(sep)
    lines.append("")

    lines.append(f"匹配文件数: {len(common_files)}")
    lines.append(f"错误总数:   {error_count}")
    if only_ocr:
        lines.append(f"仅 OCR 中有（标签无）: {len(only_ocr)} 个 — {', '.join(only_ocr)}")
    if only_gt:
        lines.append(f"仅标签中有（OCR 无）: {len(only_gt)} 个 — {', '.join(only_gt)}")
    lines.append("")

    # ── 错误明细 ─────────────────────────────────────────────
    if errors:
        lines.append(sep)
        lines.append(f"  错误明细（共 {error_count} 条）")
        lines.append(sep)
        for e in errors:
            lines.append(f"  {e['file_name']}  [{e['field']}]")
            lines.append(f"    标签: {e['label']}")
            lines.append(f"    OCR : {e['ocr']}")
            lines.append("")
    else:
        lines.append("所有字段全部正确，无错误。")

    # ── 逐文件详情 ───────────────────────────────────────────
    lines.append(sep)
    lines.append(f"  逐文件对比详情")
    lines.append(sep)
    for file_name in common_files:
        gt_row = gt_data[file_name]
        ocr_row = ocr_data[file_name]
        file_errors = []
        for col_idx, field_name in COMPARE_FIELDS.items():
            gt_val = gt_row.get(col_idx, "")
            ocr_val = ocr_row.get(col_idx, "")

            if field_name == "birth_date":
                gt_norm = _normalize_date(gt_val)
                ocr_norm = _normalize_date(ocr_val)
            elif field_name == "gender":
                gt_norm = _normalize_gender(gt_val)
                ocr_norm = _normalize_gender(ocr_val)
            else:
                gt_norm = gt_val.strip()
                ocr_norm = ocr_val.strip()

            if gt_norm != ocr_norm:
                file_errors.append(f"    {field_name}: 标签={gt_val}  OCR={ocr_val}")

        if file_errors:
            lines.append(f"  ✗ {file_name}")
            lines.extend(file_errors)
        else:
            lines.append(f"  ✓ {file_name}")
    lines.append("")

    # ── 写入 & 打印 ──────────────────────────────────────────
    report_text = "\n".join(lines)
    REPORT_TXT.write_text(report_text, encoding="utf-8")

    # 终端同步输出（GBK 不支持的字符替换为 ASCII）
    try:
        print(report_text)
    except UnicodeEncodeError:
        print(report_text.encode("gbk", errors="replace").decode("gbk"))
    print(f"评估报告已保存至: {REPORT_TXT}")


if __name__ == "__main__":
    main()
